"""
Generate the IESTP Paucartambo attendance report Excel,
matching the official format_de_asistencia.xlsx template exactly.
"""
import calendar, datetime, io
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_TEXT

# ── colour palette (from the template) ────────────────────────────────────────
NAVY        = "0B2161"   # dark navy header
NAVY2       = "17375E"   # secondary navy
GOLD        = "C0A000"   # gold/dark yellow accent
RED_HDR     = "FF0000"   # red title text
BLUE_TITLE  = "1F3864"   # dark blue title
LIGHT_BLUE  = "DCE6F1"   # column header bg
YELLOW_BG   = "FFFF00"   # T column
RED_BG      = "FF0000"   # F column
GREEN_BG    = "00B050"   # P column
ORANGE_BG   = "FFC000"   # P/S column
PURPLE_BG   = "7030A0"   # C/S column
GREY_BG     = "D9D9D9"   # Dias trabajados column
WHITE       = "FFFFFF"
LIGHT_ROW   = "EAF0FB"   # alternating row

def _side(color="000000", style="thin"):
    return Side(border_style=style, color=color)

def _border(all_sides="thin", color="000000"):
    s = _side(color, all_sides)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=10, color="000000", name="Arial"):
    return Font(bold=bold, size=size, color=color, name=name)

def _align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def build_report(students_data: list[dict], month: int, year: int) -> bytes:
    """
    students_data: list of dicts:
        {
          'number': int,
          'dni': str,
          'name': str,
          'position': str,
          'condition': str,       # Contratado / Nombrado
          'attendance': set[int], # set of day numbers (1-31) they attended
        }
    Returns bytes of the .xlsx file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    month_name_es = {
        1:"ENERO",2:"FEBRERO",3:"MARZO",4:"ABRIL",
        5:"MAYO",6:"JUNIO",7:"JULIO",8:"AGOSTO",
        9:"SETIEMBRE",10:"OCTUBRE",11:"NOVIEMBRE",12:"DICIEMBRE"
    }[month]

    days_in_month = calendar.monthrange(year, month)[1]
    # Day of week for each day (0=Mon..6=Sun)
    day_abbr_es = {0:"L",1:"M",2:"M",3:"J",4:"V",5:"S",6:"D"}

    # ── Column layout ─────────────────────────────────────────────────────────
    # A=row_num, B=DNI, C=Name, D=Position, E=Condition,
    # F..F+days-1 = day columns,
    # then T, F, P, P/S, C/S, Dias_trabajados
    COL_NUM   = 1   # A
    COL_DNI   = 2   # B
    COL_NAME  = 3   # C
    COL_CARGO = 4   # D
    COL_COND  = 5   # E
    COL_DAY1  = 6   # F  (first day column)
    COL_DAYN  = COL_DAY1 + days_in_month - 1
    COL_T     = COL_DAYN + 1
    COL_F     = COL_DAYN + 2
    COL_P     = COL_DAYN + 3
    COL_PS    = COL_DAYN + 4
    COL_CS    = COL_DAYN + 5
    COL_DIAS  = COL_DAYN + 6
    TOTAL_COLS = COL_DIAS

    # ── Row layout ────────────────────────────────────────────────────────────
    ROW_LOGO        = 1
    ROW_TITLE1      = 2
    ROW_TITLE2      = 3
    ROW_BLANK1      = 4
    ROW_SUBTITLE    = 5  # "FORMATO: REPORTE..."  (row 6 in template = row 5 here, 0-indexed offset varies)
    ROW_BLANK2      = 6
    ROW_BLANK3      = 7
    ROW_META        = 8  # DRE / PERIODO / TURNO
    ROW_BLANK4      = 9
    ROW_HDR_LABELS  = 10 # N° DNI Apellidos...  DIAS CALENDARIO
    ROW_HDR_DAYS    = 11 # day numbers 1..31
    ROW_HDR_ABBR    = 12 # D L M M J V S ...
    ROW_DATA_START  = 13

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(COL_NUM)].width   = 4
    ws.column_dimensions[get_column_letter(COL_DNI)].width   = 10
    ws.column_dimensions[get_column_letter(COL_NAME)].width  = 30
    ws.column_dimensions[get_column_letter(COL_CARGO)].width = 22
    ws.column_dimensions[get_column_letter(COL_COND)].width  = 12
    for d in range(days_in_month):
        ws.column_dimensions[get_column_letter(COL_DAY1+d)].width = 2.5
    for col in [COL_T, COL_F, COL_P, COL_PS, COL_CS, COL_DIAS]:
        ws.column_dimensions[get_column_letter(col)].width = 5

    # ── Row heights ───────────────────────────────────────────────────────────
    ws.row_dimensions[ROW_LOGO].height      = 18
    ws.row_dimensions[ROW_TITLE1].height    = 28
    ws.row_dimensions[ROW_TITLE2].height    = 36
    ws.row_dimensions[ROW_SUBTITLE].height  = 16
    ws.row_dimensions[ROW_META].height      = 20
    ws.row_dimensions[ROW_HDR_LABELS].height= 36
    ws.row_dimensions[ROW_HDR_DAYS].height  = 16
    ws.row_dimensions[ROW_HDR_ABBR].height  = 14

    # ── Helper: merge + write ─────────────────────────────────────────────────
    def mw(r1, c1, r2, c2, value, font=None, fill=None, align=None, border=None, num_fmt=None):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        cell = ws.cell(row=r1, column=c1, value=value)
        if font:   cell.font      = font
        if fill:   cell.fill      = fill
        if align:  cell.alignment = align
        if border: cell.border    = border
        if num_fmt: cell.number_format = num_fmt
        return cell

    def w(r, c, value, font=None, fill=None, align=None, border=None, num_fmt=None):
        cell = ws.cell(row=r, column=c, value=value)
        if font:   cell.font      = font
        if fill:   cell.fill      = fill
        if align:  cell.alignment = align
        if border: cell.border    = border
        if num_fmt: cell.number_format = num_fmt
        return cell

    # ── ROW 1: logo placeholder ───────────────────────────────────────────────
    # (logo image would be inserted separately — leave A1:A5 for it)
    ws.merge_cells(start_row=1, start_column=1, end_row=5, end_column=1)
    ws.cell(row=1, column=1).alignment = _align()

    # ── ROWS 2-3: Title ───────────────────────────────────────────────────────
    mw(ROW_TITLE1, 2, ROW_TITLE1, TOTAL_COLS,
       "Instituto de Educación Superior Tecnológico Público",
       font=Font(bold=True, size=16, color=RED_HDR, name="Arial"),
       align=_align())

    mw(ROW_TITLE2, 2, ROW_TITLE2, TOTAL_COLS,
       '"PAUCARTAMBO"',
       font=Font(bold=True, size=22, color=RED_HDR, name="Arial"),
       align=_align())

    # ── ROW 5: Subtitle ───────────────────────────────────────────────────────
    mw(ROW_SUBTITLE, 2, ROW_SUBTITLE, TOTAL_COLS,
       f"FORMATO: REPORTE DE ASISTENCIA  DETALLADO  PERSONAL",
       font=Font(bold=True, size=11, color="000000", name="Arial"),
       align=_align())

    # ── ROW 8: Meta row (DRE / PERIODO / TURNO) ───────────────────────────────
    mw(ROW_META, COL_NUM, ROW_META, COL_DNI,
       "DRE: PASCO",
       font=Font(bold=True, size=10, color="0070C0", name="Arial"),
       align=_align("left"))

    periodo_text = f"PERIODO(mes/año) {month_name_es}/{year}"
    mw(ROW_META, COL_NAME, ROW_META, COL_DAYN,
       periodo_text,
       font=Font(bold=True, size=10, name="Arial"),
       align=_align("center"))

    mw(ROW_META, COL_T, ROW_META, TOTAL_COLS,
       "Turno: Tarde",
       font=Font(bold=True, size=10, name="Arial"),
       align=_align("right"))

    # ── ROW 10-12: Column headers ─────────────────────────────────────────────
    hdr_fill  = _fill(LIGHT_BLUE)
    hdr_font  = Font(bold=True, size=9, name="Arial")
    hdr_align = _align("center")
    hdr_bord  = _border()

    # Static headers N°, DNI, Apellidos y Nombres, Cargo, Condicion
    static_hdrs = [
        (COL_NUM,   "N°"),
        (COL_DNI,   "DNI"),
        (COL_NAME,  "Apellidos y Nombres"),
        (COL_CARGO, "Cargo"),
        (COL_COND,  "Condicion"),
    ]
    for col, label in static_hdrs:
        ws.merge_cells(start_row=ROW_HDR_LABELS, start_column=col,
                       end_row=ROW_HDR_ABBR, end_column=col)
        cell = ws.cell(row=ROW_HDR_LABELS, column=col, value=label)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = hdr_bord

    # "DIAS CALENDARIO-{MES}" spanning all day columns
    mw(ROW_HDR_LABELS, COL_DAY1, ROW_HDR_LABELS, COL_DAYN,
       f"DIAS CALENDARIO-{month_name_es}",
       font=hdr_font, fill=hdr_fill, align=hdr_align, border=hdr_bord)

    # Day numbers row (ROW_HDR_DAYS)
    for d in range(1, days_in_month+1):
        col = COL_DAY1 + d - 1
        cell = ws.cell(row=ROW_HDR_DAYS, column=col, value=d)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = hdr_bord

    # Day abbreviation row (ROW_HDR_ABBR)
    for d in range(1, days_in_month+1):
        col = COL_DAY1 + d - 1
        weekday = datetime.date(year, month, d).weekday()
        abbr = day_abbr_es[weekday]
        cell = ws.cell(row=ROW_HDR_ABBR, column=col, value=abbr)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = hdr_bord

    # Summary column headers (T, F, P, P/S, C/S, Dias trabajados)
    summary_cols = [
        (COL_T,    "T",    YELLOW_BG,  "000000"),
        (COL_F,    "F",    RED_BG,     WHITE),
        (COL_P,    "P",    GREEN_BG,   WHITE),
        (COL_PS,   "P/S",  ORANGE_BG,  "000000"),
        (COL_CS,   "C/S",  PURPLE_BG,  WHITE),
        (COL_DIAS, "Dias\ntrabajados", GREY_BG, "000000"),
    ]
    for col, label, bg, fc in summary_cols:
        ws.merge_cells(start_row=ROW_HDR_LABELS, start_column=col,
                       end_row=ROW_HDR_ABBR, end_column=col)
        cell = ws.cell(row=ROW_HDR_LABELS, column=col, value=label)
        cell.font      = Font(bold=True, size=8, color=fc, name="Arial")
        cell.fill      = _fill(bg)
        cell.alignment = _align("center")
        cell.border    = hdr_bord

    # ── DATA ROWS ─────────────────────────────────────────────────────────────
    thin_bord = _border("thin", "AAAAAA")

    for i, s in enumerate(students_data):
        row = ROW_DATA_START + i
        ws.row_dimensions[row].height = 18

        alt_fill = _fill(LIGHT_ROW) if i % 2 == 0 else _fill(WHITE)

        def dw(col, val, bold=False, align_h="center", fill=None, num_fmt=None):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font      = Font(bold=bold, size=9, name="Arial")
            cell.fill      = fill or alt_fill
            cell.alignment = _align(align_h)
            cell.border    = thin_bord
            if num_fmt: cell.number_format = num_fmt
            return cell

        dw(COL_NUM,   s.get("number", i+1), bold=True)
        # DNI as text to preserve leading zeros
        dni_cell = dw(COL_DNI, s.get("dni", ""), num_fmt=FORMAT_TEXT)
        dni_cell.number_format = "@"
        dw(COL_NAME,  s.get("name", ""),     align_h="left")
        dw(COL_CARGO, s.get("position", ""), align_h="left")
        dw(COL_COND,  s.get("condition", "Contratado"))

        # Attendance marks
        attended_days  = s.get("attendance", set())
        attended_count = 0
        falta_count    = 0
        for d in range(1, days_in_month+1):
            col     = COL_DAY1 + d - 1
            weekday = datetime.date(year, month, d).weekday()
            is_weekend = weekday >= 5  # Sat=5, Sun=6

            if is_weekend:
                # Weekend: grey background, no mark
                cell = ws.cell(row=row, column=col, value="")
                cell.fill      = _fill("F2F2F2")
                cell.border    = thin_bord
                cell.alignment = _align()
            elif d in attended_days:
                cell = ws.cell(row=row, column=col, value="A")
                cell.font      = Font(bold=True, size=8, color="00B050", name="Arial")
                cell.fill      = alt_fill
                cell.border    = thin_bord
                cell.alignment = _align()
                attended_count += 1
            else:
                cell = ws.cell(row=row, column=col, value="F")
                cell.font      = Font(bold=True, size=8, color="FF0000", name="Arial")
                cell.fill      = alt_fill
                cell.border    = thin_bord
                cell.alignment = _align()
                falta_count += 1

        # Summary columns: T=total attended, F=faltas, P/P/S/C/S=0
        dw(COL_T,    attended_count, fill=_fill("FFFF99"))
        dw(COL_F,    falta_count,    fill=_fill("FFCCCC"))
        dw(COL_P,    0,              fill=_fill("CCFFCC"))
        dw(COL_PS,   0,              fill=_fill("FFE0A0"))
        dw(COL_CS,   0,              fill=_fill("E0CCFF"))
        dw(COL_DIAS, attended_count, fill=_fill(GREY_BG), bold=True)

    # ── FOOTER ────────────────────────────────────────────────────────────────
    last_data_row = ROW_DATA_START + len(students_data) - 1
    footer_row    = last_data_row + 2

    today = datetime.date.today()
    mw(footer_row, COL_DIAS - 2, footer_row, TOTAL_COLS,
       f"Paucartambo {today.day:02d} de {month_name_es.lower()} {year}.",
       font=Font(italic=True, size=9, name="Arial"),
       align=_align("right"))

    # ── LEYENDA ───────────────────────────────────────────────────────────────
    ley_row = footer_row + 2
    ws.cell(row=ley_row, column=COL_NUM, value="LEYENDA").font = Font(bold=True, size=9, name="Arial")

    leyenda = [
        ("A",   "Día laborado",              "L/SIN", "Licencia sindical"),
        ("T",   "Tardanza",                  "LSG",   "Licencia sin goce"),
        ("P",   "Permiso",                   "P/V",   "Permiso a cuenta de vacaciones"),
        ("C/S", "Comisión de Servicios",     "PGH",   "Permiso con goce de haber"),
        ("NTC", "No tiene clase",            "P/C",   "Permiso por compensación"),
        ("F",   "Faltó",                     "DNL",   "Dia no laborable"),
        ("P/S", "Permiso por salud",         "V",     "Vacaciones"),
        ("O",   "Onomástico",                "E",     "Evasión"),
        ("L/S", "Licencia por Salud",        "PAR",   "Paro"),
        ("LCG", "Licencia con goce",         "H",     "Huelga"),
    ]
    for j, (k1,v1,k2,v2) in enumerate(leyenda):
        r = ley_row + 1 + j
        for col, val in [(COL_NUM, k1),(COL_DNI, v1),(COL_CARGO, k2),(COL_COND, v2)]:
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = Font(size=8, name="Arial",
                             bold=(col in (COL_NUM, COL_CARGO)))
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # ── freeze panes ─────────────────────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=ROW_DATA_START, column=COL_DAY1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()