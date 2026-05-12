"""
===============================================================================
GENERADOR PROFESIONAL DE REPORTES DE ASISTENCIA
===============================================================================

Lee un CSV de marcaciones biométricas y rellena automáticamente
una plantilla Excel institucional respetando:

✓ Estilos
✓ Colores
✓ Fórmulas
✓ Bordes
✓ Formato original
✓ Estructura institucional

-------------------------------------------------------------------------------
USO:
-------------------------------------------------------------------------------

python app1/report_template.py marcaciones.csv plantilla.xlsx salida.xlsx

EJEMPLO:

python app1/report_template.py marcaciones.csv plantilla_mayo.xlsx reporte.xlsx

-------------------------------------------------------------------------------
REQUISITOS:
-------------------------------------------------------------------------------

pip install pandas openpyxl

===============================================================================
"""

import sys
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# =============================================================================
# CONFIGURACIÓN
# =============================================================================

# Columnas Excel (1-based)
COL_DNI = 2              # B
COL_DIA_INICIO = 6       # F
COL_DIA_FIN = 36         # AJ

COL_T = 37               # AK
COL_F = 38               # AL
COL_P = 39               # AM
COL_PS = 40              # AN
COL_CS = 41              # AO
COL_TRABAJADOS = 42      # AP

# Filas
FILA_PERIODO = 8
FILA_DIAS = 11
FILA_INICIO_DATOS = 13

# Colores
COLOR_ASISTIO = "008000"
COLOR_FALTO = "FF0000"

# Meses
MESES = {
    'ENERO': 1,
    'FEBRERO': 2,
    'MARZO': 3,
    'ABRIL': 4,
    'MAYO': 5,
    'JUNIO': 6,
    'JULIO': 7,
    'AGOSTO': 8,
    'SETIEMBRE': 9,
    'SEPTIEMBRE': 9,
    'OCTUBRE': 10,
    'NOVIEMBRE': 11,
    'DICIEMBRE': 12
}


# =============================================================================
# UTILIDADES
# =============================================================================

def imprimir(titulo, mensaje=""):
    print(f"\n📌 {titulo}")
    if mensaje:
        print(f"   → {mensaje}")


def col_letter(col_num):
    return get_column_letter(col_num)


# =============================================================================
# LECTURA CSV
# =============================================================================

def leer_csv(ruta_csv):
    """
    Lee CSV biométrico y devuelve:
    set((dni, año, mes, dia))
    """

    imprimir("Leyendo CSV biométrico")

    ruta_csv = Path(ruta_csv)

    if not ruta_csv.exists():
        raise FileNotFoundError(f"No existe el CSV: {ruta_csv}")

    try:
        # Leer CSV sin asumir nombres exactos
        df = pd.read_csv(
            ruta_csv,
            encoding='utf-8-sig',
            header=None
        )

    except Exception as e:
        raise Exception(f"Error leyendo CSV: {e}")

    if len(df.columns) < 2:
        raise Exception("El CSV debe tener mínimo 2 columnas")

    # Solo primeras 2 columnas
    df = df.iloc[:, :2]

    df.columns = ['dni', 'datetime']

    # Limpiar DNI
    df['dni'] = (
        df['dni']
        .astype(str)
        .str.replace(r'\D', '', regex=True)
    )

    # Eliminar vacíos
    df = df[df['dni'] != '']

    # Convertir a entero
    df['dni'] = pd.to_numeric(df['dni'], errors='coerce')

    # Eliminar inválidos
    df = df.dropna(subset=['dni'])

    # Ignorar DNIs falsos
    df = df[df['dni'] > 100]

    # Parsear fechas
    try:
        df['fecha'] = pd.to_datetime(
            df['datetime'],
            errors='coerce',
            format='mixed'
        )
    except:
        df['fecha'] = pd.to_datetime(
            df['datetime'],
            errors='coerce'
        )

    # Eliminar fechas inválidas
    df = df.dropna(subset=['fecha'])

    # Extraer componentes
    df['anio'] = df['fecha'].dt.year
    df['mes'] = df['fecha'].dt.month
    df['dia'] = df['fecha'].dt.day

    asistencias = set(
        zip(
            df['dni'].astype(int),
            df['anio'],
            df['mes'],
            df['dia']
        )
    )

    imprimir(
        "CSV procesado",
        f"{len(asistencias)} asistencias únicas detectadas"
    )

    return asistencias


# =============================================================================
# EXCEL
# =============================================================================

def extraer_periodo(ws):
    """
    Busca:
    PERIODO(mes/año) MAYO/2026

    Devuelve:
    (mes, año)
    """

    imprimir("Detectando periodo Excel")

    for cell in ws[FILA_PERIODO]:

        valor = str(cell.value or "").upper()

        match = re.search(r'([A-ZÁÉÍÓÚ]+)/(\d{4})', valor)

        if match:

            mes_texto = match.group(1).strip()
            anio = int(match.group(2))

            mes = MESES.get(mes_texto)

            if mes:
                imprimir(
                    "Periodo detectado",
                    f"{mes_texto}/{anio}"
                )
                return mes, anio

    raise Exception(
        "No se encontró el periodo en la fila 8.\n"
        "Debe existir algo como: MAYO/2026"
    )


def detectar_filas_empleados(ws):
    """
    Busca filas de empleados.
    """

    filas = []

    for fila in range(FILA_INICIO_DATOS, ws.max_row + 1):

        valor = ws.cell(row=fila, column=1).value

        if isinstance(valor, (int, float)) and valor > 0:
            filas.append(fila)

    return filas


def aplicar_estilo_asistencia(celda):
    """
    Estilo verde asistencia.
    """

    celda.font = Font(
        bold=True,
        color=COLOR_ASISTIO
    )

    celda.alignment = Alignment(
        horizontal='center',
        vertical='center'
    )


def aplicar_estilo_falta(celda):
    """
    Estilo rojo falta.
    """

    celda.font = Font(
        bold=True,
        color=COLOR_FALTO
    )

    celda.alignment = Alignment(
        horizontal='center',
        vertical='center'
    )


# =============================================================================
# PROCESAMIENTO PRINCIPAL
# =============================================================================

def procesar(ruta_csv, ruta_plantilla, ruta_salida):

    # -------------------------------------------------------------------------
    # Leer asistencias
    # -------------------------------------------------------------------------

    asistencias = leer_csv(ruta_csv)

    # -------------------------------------------------------------------------
    # Abrir Excel
    # -------------------------------------------------------------------------

    imprimir("Cargando plantilla Excel")

    ruta_plantilla = Path(ruta_plantilla)

    if not ruta_plantilla.exists():
        raise FileNotFoundError(
            f"No existe la plantilla: {ruta_plantilla}"
        )

    wb = load_workbook(ruta_plantilla)

    # Usar hoja activa
    ws = wb.active

    # -------------------------------------------------------------------------
    # Detectar periodo
    # -------------------------------------------------------------------------

    mes_num, anio = extraer_periodo(ws)

    # Filtrar asistencias del mes
    asist_mes = {
        (dni, dia)
        for (dni, a, m, dia) in asistencias
        if a == anio and m == mes_num
    }

    imprimir(
        "Asistencias filtradas",
        f"{len(asist_mes)} registros del periodo"
    )

    # -------------------------------------------------------------------------
    # Buscar empleados
    # -------------------------------------------------------------------------

    filas_empleados = detectar_filas_empleados(ws)

    imprimir(
        "Empleados detectados",
        f"{len(filas_empleados)} empleados"
    )

    # -------------------------------------------------------------------------
    # Procesar empleados
    # -------------------------------------------------------------------------

    empleados_actualizados = 0

    for fila in filas_empleados:

        dni_raw = ws.cell(
            row=fila,
            column=COL_DNI
        ).value

        if dni_raw is None:
            continue

        try:
            dni_int = int(
                str(dni_raw)
                .replace(" ", "")
                .replace("-", "")
            )

        except:
            continue

        asistio_total = 0
        faltas_total = 0

        # ---------------------------------------------------------------------
        # Recorrer días
        # ---------------------------------------------------------------------

        for col in range(COL_DIA_INICIO, COL_DIA_FIN + 1):

            celda = ws.cell(row=fila, column=col)

            dia_num = ws.cell(
                row=FILA_DIAS,
                column=col
            ).value

            # Día inválido
            if dia_num is None:
                continue

            # Fin de semana / celda vacía
            if celda.value is None:
                continue

            # Asistencia
            if (dni_int, dia_num) in asist_mes:

                celda.value = "A"

                aplicar_estilo_asistencia(celda)

                asistio_total += 1

            # Falta
            else:

                celda.value = "F"

                aplicar_estilo_falta(celda)

                faltas_total += 1

        # ---------------------------------------------------------------------
        # Fórmulas resumen
        # ---------------------------------------------------------------------

        rango = f"F{fila}:{col_letter(COL_DIA_FIN)}{fila}"

        ws.cell(
            row=fila,
            column=COL_T
        ).value = f'=COUNTIF({rango},"T")'

        ws.cell(
            row=fila,
            column=COL_F
        ).value = f'=COUNTIF({rango},"F")'

        ws.cell(
            row=fila,
            column=COL_P
        ).value = f'=COUNTIF({rango},"P")'

        ws.cell(
            row=fila,
            column=COL_PS
        ).value = f'=COUNTIF({rango},"P/S")'

        ws.cell(
            row=fila,
            column=COL_CS
        ).value = f'=COUNTIF({rango},"C/S")'

        ws.cell(
            row=fila,
            column=COL_TRABAJADOS
        ).value = f'=COUNTIF({rango},"A")'

        empleados_actualizados += 1

    # -------------------------------------------------------------------------
    # Guardar archivo
    # -------------------------------------------------------------------------

    ruta_salida = Path(ruta_salida)

    ruta_salida.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    wb.save(ruta_salida)

    imprimir(
        "Reporte generado correctamente",
        str(ruta_salida)
    )

    imprimir(
        "Resumen final",
        f"{empleados_actualizados} empleados actualizados"
    )


# =============================================================================
# MAIN
# =============================================================================

if __name__ == '__main__':

    if len(sys.argv) != 4:

        print("\nUSO:")
        print(
            "python report_template.py "
            "<csv> <plantilla.xlsx> <salida.xlsx>"
        )

        sys.exit(1)

    ruta_csv = sys.argv[1]
    ruta_plantilla = sys.argv[2]
    ruta_salida = sys.argv[3]

    try:

        procesar(
            ruta_csv,
            ruta_plantilla,
            ruta_salida
        )

        print("\n✅ PROCESO COMPLETADO")

    except Exception as e:

        print(f"\n❌ ERROR: {e}")

        sys.exit(1)